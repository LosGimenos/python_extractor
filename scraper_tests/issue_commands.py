import datetime
import pytz
import os
from selenium import webdriver
import logging
from .amazon import get_amazon_reviews
# from .walmart import get_walmart_reviews
from .walmart_soup import get_walmart_reviews
from .target import get_target_reviews_js
from .walgreens import get_walgreens_reviews_js
from .ulta import get_ulta_reviews_js
from .makeupalley import get_makeup_alley_reviews_js
from .target_from_api import get_target_reviews_api
from .models import ProductPageUrl, Product, Review, Project
from .redis_testing_queue import get_redis_queue, set_redis_queue
from openpyxl import Workbook, load_workbook
from .helpers import *

logger = logging.getLogger(__name__)

# March 21st 2016 through Oct 1, 2017.

def run_review_scraping(source_list, cutoff_date_text=None, start_date_text=None):

    if start_date_text == None:
        start_date = datetime.datetime.today().replace(tzinfo=pytz.UTC)
    else:
        start_date = datetime.datetime.strptime(start_date_text, "%B %d, %Y").replace(tzinfo=pytz.UTC)

    if cutoff_date_text == None:
        cutoff_date_text = "July 1, 2016"
    cutoff_date = datetime.datetime.strptime(cutoff_date_text, "%B %d, %Y").replace(tzinfo=pytz.UTC)

    project_name = 'Neutrogena Beauty'

    # 2 instance of block
    if Project.objects.filter(project_id=project_name).exists():
        project = Project.objects.get(project_id=project_name)
    else:
        project = Project(project_id = project_name)
        project.save()

    # source = 'MakeupAlley'
    # product_name_inclusions = ["Men's Rogaine"]
    # product_name_exclusions = ["Men's Rogaine"]

    for source in source_list:

        if source == "MakeupAlley":

            chromedriver = "/usr/local/chromedriver"
            os.environ["webdriver.chrome.driver"] = chromedriver

            # setup driver

            driver = webdriver.Chrome(chromedriver)
            driver.set_window_size(1120, 550)
            driver.set_window_position(-10000,0)

            # connect to original review url

            succeeded = False
            num_attempts = 0
            max_num_attempts = 20
            while succeeded == False:
                num_attempts += 1
                if num_attempts > max_num_attempts:
                    break
                try:
                    print("Trying")
                    driver.get("https://www.makeupalley.com/")
                    succeeded = True
                    print("Succeeded")
                except:
                    print("Error")
                    time.sleep(random.randint(2,3))

            # login

            xpath = '//li[@class="login"]/a'
            click_result = click_element_by_hitting_enter(driver,xpath)
            if click_result == "Error":
                print("Error clicking login link")

            time.sleep(2)
            username = driver.find_element_by_id("UserName")
            password = driver.find_element_by_id("Password")
            username.send_keys("asfbapp")
            password.send_keys("1asfbapp1")
            login_attempt = driver.find_element_by_xpath('//input[@id="login"]')
            login_attempt.submit()
            time.sleep(2)
            logger.warn("Logged in")

        else:

            driver = None

        for ppu in ProductPageUrl.objects.filter(project=project,source=source):

            #for testing
            # if ppu.id <= get_redis_queue(source):
            #     continue
            #end testing

            if ppu.id <= 2478:
                continue

            brand = ppu.brand
            brand_website_name = ppu.brand
            group = ppu.group
            product_line = ppu.product_line
            product = ppu.product

            product = Product(project=project,brand=brand,group=group,product_line=product_line,
                              product=product,source=source)
            product.save()

            product_page_url = ppu.url
            source_data_path = ''
            # if source == 'Amazon' or source == 'Walmart':
            #     source_data_path = ''
            # else:
            #     source_data_path = 'files/Reviews/' + brand_name + '/' + project_name + '/' + source + '/Source Data/' + ppu.source_data_path + '.txt'

            # source_data_path = 'files/Reviews/Makeup Alley/Source Data/Tanda Zap.txt'

            get_reviews(project,product,product_page_url,source_data_path,source,start_date,cutoff_date,brand_website_name,driver)

            #for testing send id and source data to redis for storage
            # set_redis_queue(ppu.id, source)
            #end testing

        if source == "MakeupAlley":

            xpath = '//a[@class="dropdown-toggle"]'
            click_result = click_element_by_hitting_enter(driver,xpath)
            if click_result == "Error":
                print("Error clicking username dropdown")

            xpath = '//a[@href="/account/logout.asp"]'
            click_result = click_element_by_hitting_enter(driver,xpath)
            if click_result == "Error":
                print("Error clicking logout")

            time.sleep(2)

            driver.quit()

def get_reviews(project,product,product_page_url,source_data_path,source,start_date,cutoff_date,brand_website_name,driver):

    # Define source list

    # source_list = []
    # source_list.append(source)

    source_list = [source]

    # Get reviews

    headers_list = []
    headers_list.append({
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36'})
    headers_list.append({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A'})
    headers_list.append({
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5355d Safari/8536.25'})
    headers_list.append({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_6_8) AppleWebKit/537.13+ (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2'})
    headers_list.append({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/534.55.3 (KHTML, like Gecko) Version/5.1.3 Safari/534.53.10'})

    if source == 'Target':
        # get_target_reviews_js(headers_list,product_page_url,project,product,start_date,cutoff_date,brand_website_name)
        get_target_reviews_api(headers_list,product_page_url,project,product,start_date,cutoff_date,brand_website_name)
    elif source == 'Walgreens':
        get_walgreens_reviews_js(product_page_url,project,product,start_date,cutoff_date,brand_website_name)
    elif source == 'Ulta':
        get_ulta_reviews_js(product_page_url,project,product,start_date,cutoff_date,brand_website_name)
    elif source == 'MakeupAlley':
        get_makeup_alley_reviews_js(product_page_url,project,product,start_date,cutoff_date,brand_website_name,driver)
    elif source == 'Amazon':
        get_amazon_reviews(headers_list,product_page_url,project,product,start_date,cutoff_date,brand_website_name)
    elif source == 'Walmart':
        get_walmart_reviews(headers_list,product_page_url,project,product,start_date,cutoff_date,brand_website_name)

def create_review_file(project, source_list, path):

    filename = project.project_id + '.xlsx'

    # Setup tabs

    wb = Workbook()
    ws_0 = wb.active
    ws_1 = wb.create_sheet("Products")
    ws_2 = wb.create_sheet("Reviews")
    wb.remove_sheet(ws_0)

    # Fill out products tab

    ws_1.cell(row = 1, column = 1).value = "Source"
    ws_1.cell(row = 1, column = 2).value = "Product Name"
    ws_1.cell(row = 1, column = 3).value = "Average Rating"
    ws_1.cell(row = 1, column = 4).value = "Total # of Reviews"
    ws_1.cell(row = 1, column = 5).value = "# of 1 star Reviews"
    ws_1.cell(row = 1, column = 6).value = "# of 2 star Reviews"
    ws_1.cell(row = 1, column = 7).value = "# of 3 star Reviews"
    ws_1.cell(row = 1, column = 8).value = "# of 4 star Reviews"
    ws_1.cell(row = 1, column = 9).value = "# of 5 star Reviews"
    ws_1.cell(row = 1, column = 10).value = "Brand"
    ws_1.cell(row = 1, column = 11).value = "Group"
    ws_1.cell(row = 1, column = 12).value = "Product Line"

    all_products = [p for p in Product.objects.filter(project=project, source__in=source_list)]

    row_index = 2

    for product in all_products:

        ws_1.cell(row = row_index, column = 1).value = product.source
        ws_1.cell(row = row_index, column = 2).value = product.product
        ws_1.cell(row = row_index, column = 3).value = product.average_rating
        ws_1.cell(row = row_index, column = 4).value = product.total_number_reviews
        ws_1.cell(row = row_index, column = 5).value = product.num_reviews_one_star
        ws_1.cell(row = row_index, column = 6).value = product.num_reviews_two_stars
        ws_1.cell(row = row_index, column = 7).value = product.num_reviews_three_stars
        ws_1.cell(row = row_index, column = 8).value = product.num_reviews_four_stars
        ws_1.cell(row = row_index, column = 9).value = product.num_reviews_five_stars
        ws_1.cell(row = row_index, column = 10).value = product.brand
        ws_1.cell(row = row_index, column = 11).value = product.group
        ws_1.cell(row = row_index, column = 12).value = product.product_line

        row_index += 1


    # Fill out reviews tab

    ws_2.cell(row = 1, column = 1).value = "Source"
    ws_2.cell(row = 1, column = 2).value = "Product Name"
    ws_2.cell(row = 1, column = 3).value = "Url"
    ws_2.cell(row = 1, column = 4).value = "Date"
    ws_2.cell(row = 1, column = 5).value = "Title"
    ws_2.cell(row = 1, column = 6).value = "Username"
    ws_2.cell(row = 1, column = 7).value = "Review Text"
    ws_2.cell(row = 1, column = 8).value = "Rating"
    ws_2.cell(row = 1, column = 9).value = "# of Comments"
    ws_2.cell(row = 1, column = 10).value = "Recommendation"
    ws_2.cell(row = 1, column = 11).value = "Is From Brand Website"
    ws_2.cell(row = 1, column = 12).value = "Skin Type"
    ws_2.cell(row = 1, column = 13).value = "Gender"
    ws_2.cell(row = 1, column = 14).value = "Brand"
    ws_2.cell(row = 1, column = 15).value = "Group"
    ws_2.cell(row = 1, column = 16).value = "Product Line"

    all_reviews = [r for r in Review.objects.filter(project=project, product__source__in=source_list)]

    row_index = 2

    for review in all_reviews:

        ws_2.cell(row = row_index, column = 1).value = review.product.source
        ws_2.cell(row = row_index, column = 2).value = review.product.product
        ws_2.cell(row = row_index, column = 3).value = review.url
        ws_2.cell(row = row_index, column = 4).value = review.date.strftime("%m/%d/%Y")
        ws_2.cell(row = row_index, column = 5).value = review.title
        ws_2.cell(row = row_index, column = 6).value = review.username
        ws_2.cell(row = row_index, column = 7).value = review.review_text
        ws_2.cell(row = row_index, column = 8).value = review.rating
        ws_2.cell(row = row_index, column = 9).value = review.num_comments
        ws_2.cell(row = row_index, column = 10).value = review.would_recommend
        ws_2.cell(row = row_index, column = 11).value = review.is_from_brand_website
        ws_2.cell(row = row_index, column = 12).value = review.skin_type
        ws_2.cell(row = row_index, column = 13).value = review.gender
        ws_2.cell(row = row_index, column = 14).value = review.product.brand
        ws_2.cell(row = row_index, column = 15).value = review.product.group
        ws_2.cell(row = row_index, column = 16).value = review.product.product_line

        row_index += 1

    # Save file

    wb.save(path + filename)

project = Project.objects.filter(id=1)[0]
print(project)

create_review_file(project, ['Amazon', 'Walmart', 'Target', 'Walgreens', 'MakeupAlley', 'Ulta'], '')
